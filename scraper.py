from bs4 import BeautifulSoup;
import regex as re;
import xlsxwriter;
from openpyxl import load_workbook;
from selenium import webdriver;
from selenium.webdriver.support.ui import WebDriverWait as wait
import dishcategorizer
import random;
import addrestaurantstosheet

# This parses menu sites written in javascript by using marionette browsers. Download gecko and place it in the same path as me.
# 3/27/2019

#########################################################################

# if you leave these blank the console will just ask you to input them. Do whichever you feel like
urlList = []
workbookName = ""


#########################################################################

##Soup functions
def getAddress(soup):
    address = ""
    for s in soup:
        if s.p:
            address = s.p.text
            # TODO: Make this more flexible
            if 'NW' in address or 'SW' in address or "Northwest" in address or "Street" in address or "ave" in address:
                return address
    try:
        address = soup.p
    except:
        print('fail')
        pass

    return 'Address not found!'


def getType(dishCategory):
    dishType = dishCategory.findAll(class_="category-name")
    return dishType


def getNames(dishCategory):
    dishNames1 = dishCategory.findAll('span', itemprop='name')
    # crams all the cleaning in one line, fuck readability
    dishNames = [dish.text.split("(")[0] for dish in dishNames1]
    return dishNames


def getDescriptions(dishCategory):
    dishDescriptions1 = dishCategory.findAll('meta', itemprop='description')
    dishDescriptions = [x['content'] for x in dishDescriptions1]
    return dishDescriptions


# Regex Functions
def makeItRegEx(aList):
    regExList = []
    for word in aList:
        regExList.append(word)
    return "|".join(regExList)


def stripText(text):
    text = " ".join(text.split(","))
    text = " ".join(text.split('.'))
    text = " ".join(text.split('('))
    text = " ".join(text.split(')'))
    text = " ".join(text.split('"'))
    text = " ".join(text.split('-'))
    text = " ".join(text.split(':')).lower()
    return text


def plantCatcher(dishDescription):
    try:
        dishDescription1 = dishDescription.string.lower().strip()
    except:
        dishDescription1 = dishDescription.lower().strip()

    dishDescription1 = stripText(dishDescription1)
    veggieWords = ["vegan", "veggie", "vegetarian", "soy", "choice of", "faux", "plant"]
    vegan = re.findall(makeItRegEx(veggieWords), dishDescription1)
    if vegan:
        return vegan[0]


def veganCatcher(dishDescription):
    try:
        dishDescription1 = dishDescription.string.lower().strip()
    except:
        dishDescription1 = dishDescription.lower().strip()

    dishDescription1 = stripText(dishDescription1)

    dairyWords = ["cheese", "cream", "butter", "cheese", "parmesan", "cheddar", "mozzarella", "ricotta", "feta"]
    eggsWords = ["egg", "omelette", "meringue", "yolk"]
    otherWords = ["honey"]
    dairy = re.findall(makeItRegEx(dairyWords), dishDescription1)
    eggs = re.findall(makeItRegEx(eggsWords), dishDescription1)
    honey = re.findall(makeItRegEx(otherWords), dishDescription1)
    if dairy:
        return dairy[0]
    elif eggs:
        return eggs[0]
    elif honey:
        return honey[0]


def meatCatcher(dishDescription):
    try:
        dishDescription1 = dishDescription.string
    except:
        dishDescription1 = dishDescription

    dishDescription1 = stripText(dishDescription1)

    chickenWords = ['chicken', 'breast', 'leg', 'thigh', 'drumstick', 'wing']
    cowWords = ['beef', 'steak', 'sausage', 'rib', 'veal']
    pigWords = ['pig', 'pork', 'ham', 'bacon', 'veal', 'prosciutto', 'pancetta', 'boar']
    seaWords = ["bass", "fish", "seafood", "clams", "mussel", "lobster", "octopus", "squid", "calamari", "scallop",
                "shrimp", "salmon", "prawn", "crab", "eel", "tuna", "mackerel", "mackrel"]
    otherWords = ["bison", "meat", "lamb", "turkey", "quail", "rabbit", "venison", "duck", "pepperoni"]

    chicken = re.findall(makeItRegEx(chickenWords), dishDescription1)
    cow = re.findall(makeItRegEx(cowWords), dishDescription1)
    pig = re.findall(makeItRegEx(pigWords), dishDescription1)
    sea = re.findall(makeItRegEx(seaWords), dishDescription1)
    other = re.findall(makeItRegEx(otherWords), dishDescription1)

    if chicken:
        return chicken[0]
    elif pig:
        return pig[0]
    elif cow:
        return cow[0]
    elif sea:
        return sea[0]
    elif other:
        return other[0]


def glutenCatcher(websiteText):
    dishDescription1 = ''.join(websiteText.findAll(text=True)).strip()
    dishDescription1 = stripText(dishDescription1)
    gluten = re.search(r'\bgluten\b', dishDescription1)
    # we now try to find the 50 chars before, after the word "gluten" to give it some context - are there really gluten free options?

    if gluten:
        n = (gluten.start())
        m = n - 50
        z = n + 150

        glutenContext = dishDescription1[m:z]
        dishDescription1.split()
        return glutenContext
    else:
        return False


# Code begins #
if workbookName == "":
    workbookName = input("Enter workbook name (including .xlsx): ")
userInput = ""
yes = "y"
if not urlList:
    while userInput.lower() != "ok":
        userInput = input("Paste a Skip restaurant URL (enter 'ok' when done): ")
        if "http" in userInput:
            urlList.append(userInput)

try:
    load_workbook(filename=workbookName)
    yes = input("This workbook already exists. Overwrite? (y/n)): ")

except:
    if yes == "y":
        pass
    else:
        raise
        input("Aborting mission! Enter any key to continue")

workbook = xlsxwriter.Workbook(workbookName)
print("Restaurant info loading...")
restaurantNames = []
for url in urlList:
    # TODO: uuughhhhh
    print("Now scraping at URL: " + url)

    browser = webdriver.Firefox(executable_path='/Users/hazelfoerstner/Desktop/Edibly/menustuff/geckodriver')
    page = browser.get(url)

    restaurantName = \
    wait(browser, 10).until(lambda browser: browser.find_element_by_xpath("//meta[@itemprop='brand']")).get_attribute(
        "content").split("(")[0]
    soup_level1 = BeautifulSoup(browser.page_source, 'lxml')
    soup = soup_level1.find_all('div')

    try:
        worksheetName = restaurantName[0:26].strip()
        worksheet = workbook.add_worksheet(worksheetName + '.xlsx')
        restaurantNames.append(restaurantName.strip())
    except:
        x = random.randint(1, 99)
        x = str(x)
        worksheetName = restaurantName[0:23] + x + '.xlsx'
        worksheet = workbook.add_worksheet(worksheetName + x + '.xlsx')

    # formatting
    bold = workbook.add_format({'bold': True})
    wrapItUp = workbook.add_format()
    wrapItUp.set_text_wrap()
    worksheet.set_column('A:A', 20)
    worksheet.set_column('B:C', 50)
    worksheet.set_column('F:G', 50)
    worksheet.set_column('G:H', 150)

    veganColor = workbook.add_format()
    veganColor.set_bg_color('lime')
    veggieColor = workbook.add_format()
    veggieColor.set_bg_color('yellow')

    # gluten = glutenCatcher(soup)
    # if gluten:
    #    worksheet.write(3, 7, gluten, wrapItUp)

    # No gluten for skip, I don't feel like figuring this out

    divs = []
    dishTypes = []

    for div in soup:
        titles = div.findAll('h6')
        if titles != [] and titles[0].text not in dishTypes:
            dishTypes.append(titles[0].text)
            divs.append(div)
    divs.pop(0)
    dishTypes.pop(0)
    row = 3
    col = 0

    n = 1
    categoryIs = 0

    for dishCategory in divs:
        dishType = dishTypes[categoryIs]
        categoryIs += 1
        dishNames = getNames(dishCategory)
        dishDescriptions = getDescriptions(dishCategory)

        # write type of dish, i.e. antipasti
        worksheet.write(row, col, dishType, wrapItUp)

        # this guy gets rid of any extra titles in the beginning, like "Pizza is fresh made in-house".
        while len(dishDescriptions) > len(dishNames):
            dishDescriptions.pop(0)
        while len(dishNames) > len(dishDescriptions):
            dishNames.pop(0)

        # this writes the dishes and dish descriptions in
        for dish in dishNames:
            worksheet.write(row, col + 1, dish, wrapItUp)
            n = dishNames.index(dish)
            # only some dishes have descriptions.
            try:
                dishDescription = dishDescriptions[n]
            except:
                dishDescription = "No description given."

            # Priority for regex:
            # non-veg names < non-veg descriptions < veg names < veg descriptions
            vegan = veganCatcher(dishDescription)
            meat = meatCatcher(dishDescription)
            plant = plantCatcher(dishDescription)
            if vegan == None:
                vegan = veganCatcher(dish)
            if meat == None:
                meat = meatCatcher(dish)
            if plant == None:
                plant = plantCatcher(dish)
            if not plant:
                worksheet.write(row, 4, meat)
                worksheet.write(row, 3, meat)
                worksheet.write(row, 3, vegan)
                worksheet.write(row, col + 2, dishDescription, wrapItUp)
            else:
                worksheet.write(row, 4, plant, veganColor)
                worksheet.write(row, 3, plant, veganColor)
                worksheet.write(row, col + 2, dishDescription, wrapItUp)
                worksheet.set_row(row, None, veganColor)

            if not meat:
                if not vegan:
                    worksheet.set_row(row, None, veganColor)
                else:
                    worksheet.set_row(row, None, veggieColor)
            row += 1

    # Write some data headers.
    worksheet.write('A1', restaurantName, bold)
    worksheet.write('B1', "Url: " + url, bold)
    worksheet.write('A2', 'Category', bold)
    worksheet.write('B2', 'Name', bold)
    worksheet.write('C2', "Description", bold)
    worksheet.write('D2', "Vegan", bold)
    worksheet.write('E2', "Veggie", bold)
    worksheet.write('F2', "Type", bold)
    worksheet.write('G2', "Restaurant Info", bold);
    worksheet.write('H2', "Gluten info", bold);
    address = getAddress(soup)
    worksheet.write('G3', address);
    worksheet.write('G4', "Edmonton, AB");
    worksheet.write('G5', "Image URL");
    worksheet.write('G6', "Restaurant tip here");
    worksheet.write('G7', "Manager or owner info here (name, phone number, email)");

    print("Excel sheet written to new sheet: " + worksheetName + "!")

workbook.close()
dishcategorizer.run(workbookName)
print("Closing workbook " + workbookName)
input("Enter any key to finish running")